import flet as ft
import re
from datetime import datetime
from tinydb import TinyDB, Query
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from threading import Timer 
from browser import document

class Cliente:
    def __init__(self, nome, telefone):
        self.nome = nome
        self.telefone = telefone
        self.servicos = []
        self.doc_id = None  # Adicionando o atributo doc_id

class Servico:
    def __init__(self, nome, valor, data):
        self.nome = nome
        self.valor = valor
        self.data = data

class BarbeariaApp:
    def __init__(self):
        self.db = TinyDB('barbearia_db.json')
        self.clientes_table = self.db.table('clientes')
        self.servicos_table = self.db.table('servicos')
        self.clientes = []
        self.cliente_atual = None
        self.carregar_dados()

    @staticmethod
    def validar_telefone(telefone):
        # Remove todos os caracteres não numéricos
        numeros = re.sub(r'\D', '', telefone)
        # Verifica se o número tem 10 ou 11 dígitos
        return len(numeros) == 10 or len(numeros) == 11

    @staticmethod
    def formatar_telefone(telefone):
        # Remove todos os caracteres não numéricos
        numeros = re.sub(r'\D', '', telefone)
        # Verifica se o número tem 11 dígitos (com DDD)
        if len(numeros) == 11:
            return f"({numeros[:2]}) {numeros[2:7]}-{numeros[7:]}"
        # Se tiver 10 dígitos, assume DDD de 2 dígitos
        elif len(numeros) == 10:
            return f"({numeros[:2]}) {numeros[2:6]}-{numeros[6:]}"
        else:
            # Se não corresponder ao formato esperado, retorna o número original
            return telefone

    def adicionar_cliente(self, e):
        nome = self.nome_input.value
        telefone = self.telefone_input.value

        # Verifica se o nome já existe
        if any(cliente.nome == nome for cliente in self.clientes):
            # Exiba o pop-up para nome duplicado
            self.page.dialog = self.dialog_nome_existente
            self.dialog_nome_existente.open = True
            self.page.update()
            return

        # Verifica se o telefone é válido
        if not self.validar_telefone(telefone):
            # Se o telefone não for válido, exiba o pop-up
            self.page.dialog = self.dialog_erro
            self.dialog_erro.open = True
            self.page.update()
            return

        # Formatando o telefone
        telefone_formatado = self.formatar_telefone(telefone)
        
        # Nome e telefone válidos, prosseguir com a adição do cliente
        novo_cliente = Cliente(nome, telefone_formatado)
        self.clientes.append(novo_cliente)
        self.salvar_cliente(novo_cliente)
        self.nome_input.value = ""
        self.telefone_input.value = ""
        self.nome_input.update()
        self.telefone_input.update()

    def atualizar_cliente(self, cliente):
        Cliente = Query()
        self.clientes_table.update({'nome': cliente.nome, 'telefone': cliente.telefone}, doc_ids=[cliente.doc_id])
        self.servicos_table.remove(Query().cliente_id == cliente.doc_id)
        for servico in cliente.servicos:
            servico_data = {
                'cliente_id': cliente.doc_id,
                'nome': servico.nome,
                'valor': servico.valor,
                'data': servico.data.isoformat()
            }
            self.servicos_table.insert(servico_data)

    def salvar_edicao_cliente(self, e):
        if self.cliente_atual:
            novo_nome = self.editar_nome_input.value
            novo_telefone = self.editar_telefone_input.value
            
            # Verifica se o nome já existe, exceto para o cliente que está sendo editado
            if any(Cliente.nome == novo_nome and Cliente != self.cliente_atual for Cliente in self.clientes):
                # Exiba o pop-up para nome duplicado
                self.page.dialog = self.dialog_nome_existente
                self.dialog_nome_existente.open = True
                self.page.update()
                return
            
             
            # Verifica se o telefone é válido
            if not self.validar_telefone(novo_telefone):
                # Se o telefone não for válido, exiba o pop-up
                self.page.dialog = self.dialog_erro
                self.dialog_erro.open = True
                self.page.update()
                return

            # Formatando o telefone
            novo_telefone_formatado = self.formatar_telefone(novo_telefone)

            if novo_nome and novo_telefone_formatado:
                self.cliente_atual.nome = novo_nome
                self.cliente_atual.telefone = novo_telefone_formatado
                self.atualizar_cliente(self.cliente_atual)
                self.telefone_cliente.value = f"Telefone: {novo_telefone_formatado}"
                self.editar_cliente_row.visible = False
                self.telefone_cliente.update()
                self.editar_cliente_row.update()

    def carregar_dados(self):
        for cliente_data in self.clientes_table.all():
            cliente = Cliente(cliente_data['nome'], cliente_data['telefone'])
            cliente.doc_id = cliente_data.doc_id  # Adicionando o doc_id ao cliente
            servicos = self.servicos_table.search(Query().cliente_id == cliente.doc_id)
            for servico_data in servicos:
                servico = Servico(servico_data['nome'], servico_data['valor'], datetime.fromisoformat(servico_data['data']))
                cliente.servicos.append(servico)
            self.clientes.append(cliente)

    def salvar_cliente(self, cliente):
        cliente_data = {'nome': cliente.nome, 'telefone': cliente.telefone}
        cliente.doc_id = self.clientes_table.insert(cliente_data)
        for servico in cliente.servicos:
            servico_data = {
                'cliente_id': cliente.doc_id,
                'nome': servico.nome,
                'valor': servico.valor,
                'data': servico.data.isoformat()
            }
            self.servicos_table.insert(servico_data)

    def buscar_cliente(self, e):
        busca = self.busca_input.value.lower() if self.busca_input.value else ''
        if not busca:
            self.resultados_busca.controls.clear()
            self.resultados_busca.update()
            return
        clientes_encontrados = [c for c in self.clientes if busca in c.nome.lower()]
        self.resultados_busca.controls.clear()
        for cliente in clientes_encontrados:
            self.resultados_busca.controls.append(
                ft.ListTile(
                    title=ft.Text(cliente.nome),
                    subtitle=ft.Text(cliente.telefone),
                    on_click=lambda e, cliente=cliente: self.selecionar_cliente_lista(cliente)
                )
            )
        self.resultados_busca.update()

    def selecionar_cliente_lista(self, cliente):
        self.cliente_atual = cliente
        if self.cliente_atual:
            self.telefone_cliente.value = f"Telefone: {self.cliente_atual.telefone}"
            self.editar_nome_input.value = self.cliente_atual.nome
            self.editar_telefone_input.value = self.cliente_atual.telefone
            self.editar_cliente_row.visible = True
        else:
            self.telefone_cliente.value = ""
            self.editar_cliente_row.visible = False
        self.telefone_cliente.update()
        self.editar_cliente_row.update()
        self.atualizar_tabela_servicos()
        self.busca_input.value = ""
        self.busca_input.update()

    def cancelar_edicao(self, e):
        self.editar_cliente_row.visible = False
        self.editar_cliente_row.update()

    def adicionar_servico(self, e):
        if self.cliente_atual:
            nome_servico = self.servico_input.value
            valor_servico = self.valor_input.value
            if nome_servico and valor_servico:
                try:
                    # Permitir entrada de valores float
                    valor = float(valor_servico.replace(",", "."))
                    novo_servico = Servico(nome_servico, valor, datetime.now())
                    self.cliente_atual.servicos.append(novo_servico)
                    self.atualizar_cliente(self.cliente_atual)
                    self.servico_input.value = ""
                    self.valor_input.value = ""
                    self.servico_input.update()
                    self.valor_input.update()
                    self.atualizar_tabela_servicos()
                except ValueError:
                    print("Valor inválido para o serviço")

    def atualizar_tabela_servicos(self):
        # Limpar as linhas existentes na tabela
        self.tabela_servicos.rows.clear()

        # Adicionar colunas extras para os botões Editar e Excluir, com padding para ajustar a posição dos títulos
        self.tabela_servicos.columns = [
            ft.DataColumn(ft.Text("Serviços/Produtos", color=ft.colors.WHITE)),
            ft.DataColumn(ft.Text("Valor", color=ft.colors.WHITE)),
            ft.DataColumn(ft.Text("Data", color=ft.colors.WHITE)),
            # Título "Editar" com padding à esquerda para empurrar à direita
            ft.DataColumn(
                ft.Container(
                    content=ft.Text("Editar", color=ft.colors.WHITE, text_align=ft.TextAlign.LEFT),
                    padding=ft.padding.only(left=20)  # Aumente o valor para mover mais para a direita
                )
            ),
            # Título "Excluir" com padding à esquerda para empurrar à direita
            ft.DataColumn(
                ft.Container(
                    content=ft.Text("Excluir", color=ft.colors.WHITE, text_align=ft.TextAlign.LEFT),
                    padding=ft.padding.only(left=20)  # Aumente o valor para mover mais para a direita
                )
            )
        ]

        if self.cliente_atual:
            for servico in self.cliente_atual.servicos:
                self.tabela_servicos.rows.append(
                    ft.DataRow(cells=[
                        ft.DataCell(ft.Text(servico.nome, color=ft.colors.WHITE)),
                        ft.DataCell(ft.Text(f"R$ {servico.valor:,.2f}".replace(",", "."), color=ft.colors.WHITE)),
                        ft.DataCell(ft.Text(servico.data.strftime("%d/%m/%Y %H:%M"), color=ft.colors.WHITE)),
                        # Centralizando o botão "Editar" dentro de um Container
                        ft.DataCell(
                            ft.Container(
                                content=ft.ElevatedButton("Editar", on_click=lambda e, s=servico: self.editar_servico(s)),
                                alignment=ft.alignment.center
                            )
                        ),
                        # Centralizando o botão "Excluir" dentro de um Container
                        ft.DataCell(
                            ft.Container(
                                content=ft.ElevatedButton("Excluir", on_click=lambda e, s=servico: self.excluir_servico(s)),
                                alignment=ft.alignment.center
                            )
                        )
                    ])
                )
        self.tabela_servicos.update()

    
    def editar_servico(self, servico):
        # Abre os campos de edição com os valores atuais do serviço
        self.servico_input.value = servico.nome
        self.valor_input.value = f"{servico.valor:,.2f}".replace(",", ".")
        self.page.update()
        
        # Atualizar o serviço ao clicar no botão "Adicionar Serviço"
        self.adicionar_servico_btn.text = "Salvar Alterações"
        self.adicionar_servico_btn.on_click = lambda e: self.salvar_edicao_servico(servico)
        self.adicionar_servico_btn.update()
    
    def salvar_edicao_servico(self, servico):
        try:
            novo_nome = self.servico_input.value
            novo_valor = float(self.valor_input.value.replace(",", "."))
            servico.nome = novo_nome
            servico.valor = novo_valor
            servico.data = datetime.now()  # Atualiza a data para o momento da edição
            
            # Atualiza o cliente no banco de dados
            self.atualizar_cliente(self.cliente_atual)
            self.atualizar_tabela_servicos()
            
            # Limpar campos e resetar botão
            self.servico_input.value = ""
            self.valor_input.value = ""
            self.adicionar_servico_btn.text = "Adicionar Serviço"
            self.adicionar_servico_btn.on_click = self.adicionar_servico
            self.adicionar_servico_btn.update()
            self.servico_input.update()
            self.valor_input.update()

        except ValueError:
            print("Valor inválido para o serviço")
    
    def excluir_servico(self, servico):
        self.cliente_atual.servicos.remove(servico)
        self.atualizar_cliente(self.cliente_atual)
        self.atualizar_tabela_servicos()


    def ocultar_mensagem(self):
        self.planilha_mensagem.visible = False
        self.planilha_mensagem.update()
    
    def ocultar_ganhos(self):
        self.total_ganhos.visible = False
        self.total_ganhos.update()
 
    def gerar_planilha(self, e):
        if self.cliente_atual:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Serviços de {self.cliente_atual.nome}"
            headers = ["Serviços/produtos", "Valor", "Data"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            total = 0
            for row, servico in enumerate(self.cliente_atual.servicos, start=2):
                ws.cell(row=row, column=1, value=servico.nome)
                ws.cell(row=row, column=2, value=f"R$ {servico.valor:,.2f}".replace(",", "."))
                ws.cell(row=row, column=3, value=servico.data.strftime("%d/%m/%Y %H:%M"))
                total += servico.valor
            # Adiciona a linha de total
            total_row = len(self.cliente_atual.servicos) + 2
            total_cell = ws.cell(row=total_row, column=1, value="Total")
            total_value_cell = ws.cell(row=total_row, column=2, value=f"R${total:,.2f}".replace(",", "."))
            total_cell.font = Font(bold=True)
            total_value_cell.font = Font(bold=True)
            for col in range(1, 4):
                ws.column_dimensions[get_column_letter(col)].auto_size = True
            filename = f"{self.cliente_atual.nome}_servicos.xlsx"
            wb.save(filename)
            self.planilha_mensagem.value = f"Planilha gerada com sucesso: {filename}"
            self.planilha_mensagem.visible = True
            self.planilha_mensagem.update()
             # Ocultar a mensagem após 3 segundos
            Timer(3.0, self.ocultar_mensagem).start()

    def calcular_ganhos(self, e):
        mes = int(self.mes_input.value) if self.mes_input.value else None
        ano = int(self.ano_input.value) if self.ano_input.value else None
        total = 0
        for cliente in self.clientes:
            for servico in cliente.servicos:
                if (mes is None or servico.data.month == mes) and (ano is None or servico.data.year == ano):
                    total += servico.valor
        periodo = ""
        if mes and ano:
            periodo = f"em {mes}/{ano}"
        elif mes:
            periodo = f"no mês {mes}"
        elif ano:
            periodo = f"no ano {ano}"
        else:
            periodo = "no total"
        self.total_ganhos.visible = True #Habilita a função novamente
        total_formatado = f"{total:,.2f}".replace(",", ".")
        self.total_ganhos.value = f"Total de ganhos {periodo}: R$ {total_formatado}"
        self.total_ganhos.update()
         # Ocultar a mensagem após 3 segundos
        Timer(3.0, self.ocultar_ganhos).start()
    
    
        
    def gerar_planilha_ganhos(self, e):
        mes = int(self.mes_input.value) if self.mes_input.value else None
        ano = int(self.ano_input.value) if self.ano_input.value else None
        wb = Workbook()
        ws = wb.active
        ws.title = "Ganhos"
        headers = ["Cliente", "Serviços/produtos", "Valor", "Data"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        row = 2
        total_geral = 0
        for cliente in self.clientes:
            for servico in cliente.servicos:
                if (mes is None or servico.data.month == mes) and (ano is None or servico.data.year == ano):
                    ws.cell(row=row, column=1, value=cliente.nome)
                    ws.cell(row=row, column=2, value=servico.nome)
                    ws.cell(row=row, column=3, value=f"R$ {servico.valor:,.2f}".replace(",", "."))
                    ws.cell(row=row, column=4, value=servico.data.strftime("%d/%m/%Y %H:%M"))
                    total_geral += servico.valor
                    row += 1
        # Adiciona a linha de total geral
        total_geral_cell = ws.cell(row=row, column=1, value="Total Geral")
        total_geral_value_cell = ws.cell(row=row, column=3, value=f"R$ {total_geral:,.2f}".replace(",", "."))
        total_geral_cell.font = Font(bold=True)
        total_geral_value_cell.font = Font(bold=True)
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
        filename = f"ganhos_{mes or 'todos'}_{ano or 'todos'}.xlsx"
        wb.save(filename)
        self.planilha_mensagem.value = f"Planilha de ganhos gerada: {filename}"
        self.planilha_mensagem.visible = True
        self.planilha_mensagem.update()
         # Ocultar a mensagem após 3 segundos
        Timer(3.0, self.ocultar_mensagem).start()
    
    # Funções para fechar pop-up de erros
    #
    # Função para fechar o dialog de erro de telefone incorreto
    def fechar_dialog_erro(self, e):
        self.dialog_erro.open = False
        self.page.update()
    
    # Função para fechar o dialog de erro de nome existente
    def fechar_dialog_nome_existente(self, e):
        self.dialog_nome_existente.open = False
        self.page.update()
        
    def main(self, page: ft.Page):
        self.page = page  # Armazena a referência à página na instância da classe
        page.title = ""
        page.theme_mode = ft.ThemeMode.DARK
        page.bgcolor = ft.colors.BLUE_GREY_900
        page.padding = 5
        
        # Adicione a imagem de fundo
        document.body.style.backgroundImage = r"C:\Users\pc\Music\Barber\venv\luxa.org-opacity-changed-FULLHD.jpg"
        document.body.style.backgroundSize = "cover"
        document.body.style.backgroundRepeat = "no-repeat"
        document.body.style.backgroundAttachment = "fixed"
        
        
        self.titulo = ft.Text("", size=32, color=ft.colors.WHITE, text_align=ft.TextAlign.CENTER)
        self.nome_input = ft.TextField(label="Nome do Cliente", expand=True, color=ft.colors.WHITE)
        self.telefone_input = ft.TextField(label="Telefone do Cliente", expand=True, on_change=self.formatar_telefone_input, color=ft.colors.WHITE)

        """Dialog Pop-up para nome existente"""
        self.dialog_nome_existente = ft.AlertDialog(
            title=ft.Text("Erro"),
            content=ft.Text("Nome do cliente já existe."),
            actions=[ft.TextButton("OK", on_click=self.fechar_dialog_nome_existente)],
            actions_alignment=ft.MainAxisAlignment.END
        )
        page.add(self.dialog_nome_existente)

        """Dialog Pop-up para telefone incorreto"""
        self.dialog_erro = ft.AlertDialog(
            title=ft.Text("Erro"),
            content=ft.Text("Número de telefone inválido."),
            actions=[ft.TextButton("OK", on_click=self.fechar_dialog_erro)],
            actions_alignment=ft.MainAxisAlignment.END
        )
        page.add(self.dialog_erro)

        self.adicionar_cliente_btn = ft.ElevatedButton("Adicionar Cliente", on_click=self.adicionar_cliente)
        self.telefone_cliente = ft.Text("", size=16, color=ft.colors.WHITE)
        self.busca_input = ft.TextField(label="Buscar Cliente", expand=1, on_change=self.buscar_cliente, color=ft.colors.WHITE)
        self.resultados_busca = ft.Column()
        self.editar_nome_input = ft.TextField(label="Novo Nome", expand=1, color=ft.colors.WHITE)
        self.editar_telefone_input = ft.TextField(label="Novo Telefone", expand=1, color=ft.colors.WHITE)
        self.salvar_edicao_btn = ft.ElevatedButton("Salvar Alterações", on_click=self.salvar_edicao_cliente)
        self.cancelar_edicao_btn = ft.ElevatedButton("Cancelar", on_click=self.cancelar_edicao)
        self.editar_cliente_row = ft.Row([self.editar_nome_input, self.editar_telefone_input, self.salvar_edicao_btn, self.cancelar_edicao_btn], visible=False)
        self.servico_input = ft.TextField(label="Serviços/Produtos", expand=1, color=ft.colors.WHITE)
        self.valor_input = ft.TextField(label="R$ ", expand=1, color=ft.colors.WHITE)
        self.adicionar_servico_btn = ft.ElevatedButton("Adicionar Serviço", on_click=self.adicionar_servico)
        self.tabela_servicos = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Serviços/Produtos", color=ft.colors.WHITE)),
                ft.DataColumn(ft.Text("Valor", color=ft.colors.WHITE)),
                ft.DataColumn(ft.Text("Data", color=ft.colors.WHITE))
            ],
            rows=[]
        )
        self.gerar_planilha_btn = ft.ElevatedButton("Gerar Planilha de Cliente", on_click=self.gerar_planilha)
        self.planilha_mensagem = ft.Text("", size=16, color=ft.colors.GREEN, visible=False)
        self.mes_input = ft.Dropdown(label="Mês", options=[ft.dropdown.Option(str(i)) for i in range(1, 13)], width=120)
        self.ano_input = ft.Dropdown(label="Ano", options=[ft.dropdown.Option(str(i)) for i in range(datetime.now().year-5, datetime.now().year+1)], width=120)
        self.calcular_ganhos_btn = ft.ElevatedButton("Calcular Ganhos", on_click=self.calcular_ganhos)
        self.total_ganhos = ft.Text("", size=20, color=ft.colors.WHITE)
        self.gerar_planilha_ganhos_btn = ft.ElevatedButton("Gerar Planilha de Ganhos", on_click=self.gerar_planilha_ganhos)

        # Layout
        coluna_esquerda = ft.Column(
            [   ft.Container(height=50),  # Adicionando espaço extra no topo
                ft.Row(
                    [
                        self.nome_input, self.telefone_input, self.adicionar_cliente_btn
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                ),
                ft.Container(height=20),
                ft.Row([self.busca_input], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                self.resultados_busca,
                ft.Container(height=10),
                ft.Row([self.telefone_cliente], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                self.editar_cliente_row,
                ft.Container(height=20),
                ft.Row(
                    [
                        self.servico_input, self.valor_input, self.adicionar_servico_btn
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                ),
                ft.Container(height=20),
                self.tabela_servicos,
                ft.Container(height=10),
                ft.Row([self.gerar_planilha_btn], alignment=ft.MainAxisAlignment.END),
                self.planilha_mensagem,
            ], scroll=ft.ScrollMode.AUTO, expand=True
        )

        coluna_direita = ft.Column(
            [
                ft.Text("Cálculo de Ganhos", size=24, color=ft.colors.WHITE, text_align=ft.TextAlign.CENTER),
                ft.Container(height=10),
                ft.Row([self.mes_input, self.ano_input], alignment=ft.MainAxisAlignment.CENTER),
                ft.Container(height=10),
                ft.Row([self.calcular_ganhos_btn, self.gerar_planilha_ganhos_btn], alignment=ft.MainAxisAlignment.CENTER),
                ft.Container(height=10),
                self.total_ganhos,
            ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER
        )

        conteudo_principal = ft.Column(
            [
                ft.Row([self.titulo], alignment=ft.MainAxisAlignment.CENTER),
                ft.Container(height=20),
                ft.Row(
                    [
                        coluna_esquerda,
                        ft.VerticalDivider(width=1, color=ft.colors.GREY_400),
                        coluna_direita
                    ], expand=True
                )
            ], expand=True
        )

        # Adicionando a imagem de fundo
        # def on_resize(e):
        #     background_image.width = page.width
        #     background_image.height = page.height
        #     background_image.update()

        # background_image = ft.Image(
        #     src="luxa.org-opacity-changed-FULLHD.jpg",
        #     fit=ft.ImageFit.COVER,
        #     width=page.width,
        #     height=page.height
        # )

        # background_container = ft.Container(
        #     content=ft.Stack(
        #         [
        #             background_image,
        #             conteudo_principal
        #         ]
        #     ), expand=True
        # )

        # page.on_resize = on_resize
        # page.add(background_container)

    def formatar_telefone_input(self, e):
        self.telefone_input.value = self.formatar_telefone(self.telefone_input.value)
        self.telefone_input.update()

def main(page: ft.Page):
    app = BarbeariaApp()
    app.main(page)

if __name__ == "__main__":
    ft.app(target=main)